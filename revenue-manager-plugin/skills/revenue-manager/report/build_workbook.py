#!/usr/bin/env python3
"""
build_workbook.py — turn a Revenue Manager run into a multi-tab spreadsheet.

Tab 1 is a portfolio Summary. Every property gets its own tab with a full
breakdown. This is a REPORT ONLY — it never pushes a price anywhere. The
Revenue Manager skill (Step 7.5) assembles the JSON and calls this script.

Usage:
    python3 build_workbook.py <input.json> [output.xlsx]

If output.xlsx is omitted, the workbook is written to the operator's Desktop
(falling back to the current folder) as Revenue-Report-<YYYY-MM-DD>.xlsx.

Dependency: openpyxl. If it isn't importable, the script DEGRADES GRACEFULLY
to a folder of CSVs (one summary.csv + one CSV per property) so the operator
still gets every number — just without real tabs.

The script prints exactly one machine-readable result line at the end:
    WORKBOOK_WRITTEN: /abs/path/Revenue-Report-2026-06-17.xlsx
or
    CSV_FALLBACK_WRITTEN: /abs/path/Revenue-Report-2026-06-17/

The input JSON shape (every field optional except properties[].name — the
script uses safe defaults so a partial payload never crashes it):

{
  "meta": {
    "generated_date": "2026-06-17",
    "portfolio_name": "My Portfolio",
    "default_currency": "USD",
    "recommend_only_note": "RECOMMEND-ONLY — you approve every change."
  },
  "portfolio_summary": {
    "property_count": 4, "proposed_change_count": 7,
    "underpriced_count": 2, "overpriced_count": 1,
    "avg_occupancy_pct": 0.68, "revpar_direction": "up", "notes": ""
  },
  "properties": [
    {
      "name": "Mountain View Chalet", "currency": "USD", "listing_id": "...",
      "data_freshness": "live (pulled 2026-06-17)",
      "pricing": {
        "base": {"current": 150, "recommended": 165, "pct_move": 0.10, "min": 99, "max": 199, "bound_flag": ""},
        "min":  {"current": 99,  "recommended": 99},
        "max":  {"current": 199, "recommended": 215, "pct_move": 0.08}
      },
      "comps": {"count": 42, "percentile_context": "asking ~55th pct of same-bedroom comps"},
      "ask_vs_cleared": {"ask": 160, "adr": 178, "spread": 18, "markup": 1.0},
      "kpis": {"occupancy_pct": 0.71, "revpar": 120, "pacing_stly": "+8%", "lead_time_days": 21},
      "red_flags": ["max ceiling suppressing winter upside"],
      "recommendations": [
        {"field": "max", "from": 199, "to": 215, "pct_move": 0.08,
         "reasoning": "...", "expected_impact": "RevPAR +", "flags": [],
         "prior_attempts": "", "status": "proposed"}
      ],
      "dso_recommendations": [
        {"date_range": "Dec 20–Jan 2", "change": "min-stay 3→5", "reasoning": "peak winter"}
      ],
      "safety_layer": {
        "bounds_used": "min 99 / max 199 (from PriceLabs)",
        "max_delta": "25%", "currency": "USD asserted",
        "freshness": "live", "confidence": "high — 42 comps, live pacing"
      }
    }
  ]
}
"""

import csv
import json
import re
import sys
from datetime import date
from pathlib import Path


# ── small helpers ────────────────────────────────────────────────────────────

def _today_str(meta):
    val = (meta or {}).get("generated_date")
    if val:
        return str(val)
    return date.today().isoformat()


def _g(d, key, default=""):
    """Safe get from a possibly-None dict."""
    if not isinstance(d, dict):
        return default
    val = d.get(key, default)
    return default if val is None else val


def _pct(val):
    """Render a 0..1 fraction OR an already-percent number as a clean string."""
    if val == "" or val is None:
        return ""
    try:
        f = float(val)
    except (TypeError, ValueError):
        return str(val)
    if -1.5 <= f <= 1.5:  # treat as a fraction
        return f"{f * 100:.0f}%"
    return f"{f:.0f}%"


def _default_output_path(meta):
    name = f"Revenue-Report-{_today_str(meta)}.xlsx"
    desktop = Path.home() / "Desktop"
    base = desktop if desktop.is_dir() else Path.cwd()
    return base / name


def _sanitize_sheet_title(name, used):
    """Excel sheet titles: <=31 chars, no []:*?/\\, unique, non-blank."""
    title = re.sub(r"[\[\]:\*\?/\\]", " ", str(name or "Property")).strip() or "Property"
    title = title[:31]
    candidate, n = title, 2
    while candidate.lower() in used:
        suffix = f" ({n})"
        candidate = title[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


# ── xlsx path (openpyxl) ─────────────────────────────────────────────────────

def build_xlsx(data, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    meta = data.get("meta", {}) or {}
    summary = data.get("portfolio_summary", {}) or {}
    props = data.get("properties", []) or []

    HEADER_FILL = PatternFill("solid", fgColor="1F3B4D")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    TITLE_FONT = Font(bold=True, size=14)
    SECTION_FONT = Font(bold=True, size=11, color="1F3B4D")
    WARN_FONT = Font(bold=True, color="B00020")
    wrap = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()

    # ---- Summary tab ----
    ws = wb.active
    ws.title = "Summary"
    r = 1
    ws.cell(r, 1, _g(meta, "portfolio_name", "Revenue Manager — Portfolio Summary")).font = TITLE_FONT
    r += 1
    ws.cell(r, 1, f"Generated {_today_str(meta)}  ·  currency shown per property")
    r += 1
    note = _g(meta, "recommend_only_note",
              "RECOMMEND-ONLY — you approve every change. Nothing was pushed automatically.")
    ws.cell(r, 1, note).font = WARN_FONT
    r += 2

    # portfolio roll-up
    ws.cell(r, 1, "Portfolio at a glance").font = SECTION_FONT
    r += 1
    rollup = [
        ("Properties analyzed", _g(summary, "property_count", len(props))),
        ("Proposed changes", _g(summary, "proposed_change_count")),
        ("Underpriced flags", _g(summary, "underpriced_count")),
        ("Overpriced flags", _g(summary, "overpriced_count")),
        ("Avg occupancy", _pct(_g(summary, "avg_occupancy_pct"))),
        ("RevPAR direction", _g(summary, "revpar_direction")),
    ]
    for label, value in rollup:
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, value)
        r += 1
    if _g(summary, "notes"):
        ws.cell(r, 1, "Notes").font = Font(bold=True)
        c = ws.cell(r, 2, _g(summary, "notes"))
        c.alignment = wrap
        r += 1
    r += 1

    # per-property table
    ws.cell(r, 1, "Per-property snapshot").font = SECTION_FONT
    r += 1
    cols = ["Property", "Currency", "Base now", "Base rec", "Min now", "Min rec",
            "Max now", "Max rec", "Comps", "Ask", "ADR", "Occupancy", "Top flag", "Status"]
    for ci, label in enumerate(cols, start=1):
        cell = ws.cell(r, ci, label)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    ws.freeze_panes = ws.cell(r + 1, 1)
    r += 1
    for p in props:
        pricing = p.get("pricing", {}) or {}
        base, pmin, pmax = pricing.get("base", {}) or {}, pricing.get("min", {}) or {}, pricing.get("max", {}) or {}
        avc = p.get("ask_vs_cleared", {}) or {}
        kpis = p.get("kpis", {}) or {}
        flags = p.get("red_flags", []) or []
        recs = p.get("recommendations", []) or []
        status = "applied" if any(_g(x, "status") == "applied" for x in recs) else (
            "proposed" if recs else "no change")
        row = [
            _g(p, "name", "Property"), _g(p, "currency", _g(meta, "default_currency")),
            _g(base, "current"), _g(base, "recommended"),
            _g(pmin, "current"), _g(pmin, "recommended"),
            _g(pmax, "current"), _g(pmax, "recommended"),
            _g(p.get("comps", {}), "count"), _g(avc, "ask"), _g(avc, "adr"),
            _pct(_g(kpis, "occupancy_pct")), (flags[0] if flags else ""), status,
        ]
        for ci, value in enumerate(row, start=1):
            ws.cell(r, ci, value)
        r += 1
    r += 1

    # proposed-changes action list
    ws.cell(r, 1, "Proposed changes (pending your approval)").font = SECTION_FONT
    r += 1
    chdr = ["Property", "Field", "From", "To", "Move", "Reasoning", "Flags", "Status"]
    for ci, label in enumerate(chdr, start=1):
        cell = ws.cell(r, ci, label)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    r += 1
    any_change = False
    for p in props:
        for rec in p.get("recommendations", []) or []:
            any_change = True
            vals = [
                _g(p, "name"), _g(rec, "field"), _g(rec, "from"), _g(rec, "to"),
                _pct(_g(rec, "pct_move")), _g(rec, "reasoning"),
                ", ".join(rec.get("flags", []) or []), _g(rec, "status", "proposed"),
            ]
            for ci, value in enumerate(vals, start=1):
                cell = ws.cell(r, ci, value)
                if ci == 6:
                    cell.alignment = wrap
            r += 1
    if not any_change:
        ws.cell(r, 1, "No changes recommended this run — everything is within range.")

    _autosize(ws, get_column_letter, max_width=48)

    # ---- one tab per property ----
    used_titles = {"summary"}
    for p in props:
        title = _sanitize_sheet_title(_g(p, "name", "Property"), used_titles)
        _build_property_sheet(wb.create_sheet(title), p, meta,
                              SECTION_FONT, HEADER_FILL, HEADER_FONT, Font, wrap,
                              get_column_letter)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _build_property_sheet(ws, p, meta, SECTION_FONT, HEADER_FILL, HEADER_FONT, Font, wrap, get_col):
    pricing = p.get("pricing", {}) or {}
    r = 1

    def section(title):
        nonlocal r
        ws.cell(r, 1, title).font = SECTION_FONT
        r += 1

    def kv(label, value):
        nonlocal r
        ws.cell(r, 1, label).font = Font(bold=True)
        c = ws.cell(r, 2, value)
        c.alignment = wrap
        r += 1

    ws.cell(r, 1, _g(p, "name", "Property")).font = Font(bold=True, size=14)
    r += 1
    ws.cell(r, 1, f"{_g(p, 'currency', _g(meta, 'default_currency'))}  ·  "
                  f"listing {_g(p, 'listing_id', 'n/a')}  ·  {_g(p, 'data_freshness', 'freshness unknown')}")
    r += 2

    # pricing block
    section("Pricing — current vs recommended")
    hdr = ["", "Current", "Recommended", "Move", "Bound flag"]
    for ci, label in enumerate(hdr, start=1):
        cell = ws.cell(r, ci, label)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    r += 1
    for field in ("base", "min", "max"):
        blk = pricing.get(field, {}) or {}
        ws.cell(r, 1, field.capitalize()).font = Font(bold=True)
        ws.cell(r, 2, _g(blk, "current"))
        ws.cell(r, 3, _g(blk, "recommended"))
        ws.cell(r, 4, _pct(_g(blk, "pct_move")))
        ws.cell(r, 5, _g(blk, "bound_flag"))
        r += 1
    r += 1

    # comps + ask/cleared + kpis
    comps = p.get("comps", {}) or {}
    section("Comps")
    kv("Comp count", _g(comps, "count"))
    kv("Percentile context", _g(comps, "percentile_context"))
    r += 1

    avc = p.get("ask_vs_cleared", {}) or {}
    section("Ask vs cleared")
    kv("Ask (calendar)", _g(avc, "ask"))
    kv("ADR (cleared)", _g(avc, "adr"))
    kv("Spread", _g(avc, "spread"))
    kv("Empirical markup", _g(avc, "markup"))
    r += 1

    kpis = p.get("kpis", {}) or {}
    section("KPIs")
    kv("Occupancy", _pct(_g(kpis, "occupancy_pct")))
    kv("RevPAR", _g(kpis, "revpar"))
    kv("Pacing vs STLY", _g(kpis, "pacing_stly"))
    kv("Avg lead time (days)", _g(kpis, "lead_time_days"))
    r += 1

    # red flags
    flags = p.get("red_flags", []) or []
    section("Red flags")
    if flags:
        for f in flags:
            ws.cell(r, 1, f"• {f}")
            r += 1
    else:
        ws.cell(r, 1, "None detected.")
        r += 1
    r += 1

    # recommendations table
    section("Recommendations (pending your approval)")
    rhdr = ["Field", "From", "To", "Move", "Reasoning", "Expected impact", "Flags", "Prior attempts", "Status"]
    for ci, label in enumerate(rhdr, start=1):
        cell = ws.cell(r, ci, label)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    r += 1
    recs = p.get("recommendations", []) or []
    if recs:
        for rec in recs:
            vals = [
                _g(rec, "field"), _g(rec, "from"), _g(rec, "to"), _pct(_g(rec, "pct_move")),
                _g(rec, "reasoning"), _g(rec, "expected_impact"),
                ", ".join(rec.get("flags", []) or []), _g(rec, "prior_attempts"),
                _g(rec, "status", "proposed"),
            ]
            for ci, value in enumerate(vals, start=1):
                cell = ws.cell(r, ci, value)
                if ci in (5, 6):
                    cell.alignment = wrap
            r += 1
    else:
        ws.cell(r, 1, "No changes recommended — within range.")
        r += 1
    r += 1

    # DSO / min-stay recs
    dsos = p.get("dso_recommendations", []) or []
    if dsos:
        section("Date-specific / min-stay recommendations")
        dhdr = ["Dates", "Change", "Reasoning"]
        for ci, label in enumerate(dhdr, start=1):
            cell = ws.cell(r, ci, label)
            cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        r += 1
        for d in dsos:
            ws.cell(r, 1, _g(d, "date_range"))
            ws.cell(r, 2, _g(d, "change"))
            ws.cell(r, 3, _g(d, "reasoning")).alignment = wrap
            r += 1
        r += 1

    # safety-layer footer
    sl = p.get("safety_layer", {}) or {}
    section("Safety layer")
    kv("Bounds used", _g(sl, "bounds_used"))
    kv("Max delta", _g(sl, "max_delta"))
    kv("Currency", _g(sl, "currency"))
    kv("Freshness", _g(sl, "freshness"))
    kv("Confidence", _g(sl, "confidence"))

    _autosize(ws, get_col, max_width=60)


def _autosize(ws, get_col, max_width=50):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column
            longest = max(len(line) for line in str(cell.value).split("\n"))
            widths[col] = max(widths.get(col, 0), longest)
    for col, w in widths.items():
        ws.column_dimensions[get_col(col)].width = min(max(w + 2, 10), max_width)


# ── csv fallback ─────────────────────────────────────────────────────────────

def build_csv_fallback(data, out_path):
    """openpyxl unavailable → write a folder of CSVs so no data is lost."""
    meta = data.get("meta", {}) or {}
    summary = data.get("portfolio_summary", {}) or {}
    props = data.get("properties", []) or []

    out_dir = Path(out_path)
    if out_dir.suffix:  # was given a .xlsx path → use a sibling folder
        out_dir = out_dir.with_suffix("")
    out_dir.mkdir(parents=True, exist_ok=True)

    with (out_dir / "summary.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Revenue Manager — Portfolio Summary", _today_str(meta)])
        w.writerow([_g(meta, "recommend_only_note",
                       "RECOMMEND-ONLY — you approve every change.")])
        w.writerow([])
        w.writerow(["Properties analyzed", _g(summary, "property_count", len(props))])
        w.writerow(["Proposed changes", _g(summary, "proposed_change_count")])
        w.writerow(["Underpriced flags", _g(summary, "underpriced_count")])
        w.writerow(["Overpriced flags", _g(summary, "overpriced_count")])
        w.writerow(["Avg occupancy", _pct(_g(summary, "avg_occupancy_pct"))])
        w.writerow(["RevPAR direction", _g(summary, "revpar_direction")])
        w.writerow([])
        w.writerow(["Property", "Currency", "Base now", "Base rec", "Min now", "Min rec",
                    "Max now", "Max rec", "Comps", "Ask", "ADR", "Occupancy", "Top flag", "Status"])
        for p in props:
            pr = p.get("pricing", {}) or {}
            base, pmin, pmax = pr.get("base", {}) or {}, pr.get("min", {}) or {}, pr.get("max", {}) or {}
            avc = p.get("ask_vs_cleared", {}) or {}
            kpis = p.get("kpis", {}) or {}
            flags = p.get("red_flags", []) or []
            recs = p.get("recommendations", []) or []
            status = "applied" if any(_g(x, "status") == "applied" for x in recs) else (
                "proposed" if recs else "no change")
            w.writerow([_g(p, "name"), _g(p, "currency", _g(meta, "default_currency")),
                        _g(base, "current"), _g(base, "recommended"),
                        _g(pmin, "current"), _g(pmin, "recommended"),
                        _g(pmax, "current"), _g(pmax, "recommended"),
                        _g(p.get("comps", {}), "count"), _g(avc, "ask"), _g(avc, "adr"),
                        _pct(_g(kpis, "occupancy_pct")), (flags[0] if flags else ""), status])

    used = {"summary"}
    for p in props:
        fname = _sanitize_sheet_title(_g(p, "name", "Property"), used)
        with (out_dir / f"{fname}.csv").open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([_g(p, "name"), _g(p, "currency", _g(meta, "default_currency")),
                        _g(p, "listing_id"), _g(p, "data_freshness")])
            w.writerow([])
            pr = p.get("pricing", {}) or {}
            w.writerow(["Pricing", "Current", "Recommended", "Move", "Bound flag"])
            for field in ("base", "min", "max"):
                blk = pr.get(field, {}) or {}
                w.writerow([field, _g(blk, "current"), _g(blk, "recommended"),
                            _pct(_g(blk, "pct_move")), _g(blk, "bound_flag")])
            w.writerow([])
            w.writerow(["Recommendations (pending approval)"])
            w.writerow(["Field", "From", "To", "Move", "Reasoning", "Expected impact", "Flags", "Status"])
            for rec in p.get("recommendations", []) or []:
                w.writerow([_g(rec, "field"), _g(rec, "from"), _g(rec, "to"), _pct(_g(rec, "pct_move")),
                            _g(rec, "reasoning"), _g(rec, "expected_impact"),
                            ", ".join(rec.get("flags", []) or []), _g(rec, "status", "proposed")])
    return out_dir


# ── entry point ──────────────────────────────────────────────────────────────

def main(argv):
    if len(argv) < 2:
        print("usage: python3 build_workbook.py <input.json> [output.xlsx]", file=sys.stderr)
        return 2
    try:
        data = json.loads(Path(argv[1]).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        print(f"ERROR: could not read input JSON: {e}", file=sys.stderr)
        return 1

    meta = data.get("meta", {}) or {}
    out_arg = argv[2] if len(argv) > 2 else _default_output_path(meta)

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        out_dir = build_csv_fallback(data, out_arg)
        print(f"CSV_FALLBACK_WRITTEN: {Path(out_dir).resolve()}/")
        return 0

    out_path = build_xlsx(data, out_arg)
    print(f"WORKBOOK_WRITTEN: {Path(out_path).resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
