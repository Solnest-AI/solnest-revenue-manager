#!/usr/bin/env python3
"""
Smoke test for build_workbook.py.

Always tests the CSV fallback (no deps). Tests the .xlsx path too if openpyxl
is importable. Run from anywhere:

    python3 smoke_test.py
"""

import json
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

import build_workbook as bw  # noqa: E402


def _load_sample():
    return json.loads((HERE / "sample_data.json").read_text(encoding="utf-8"))


def test_csv_fallback():
    data = _load_sample()
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "Revenue-Report-test.xlsx"
        out_dir = bw.build_csv_fallback(data, out)
        assert out_dir.is_dir(), "fallback dir not created"
        assert (out_dir / "summary.csv").exists(), "summary.csv missing"
        prop_csvs = [p.name for p in out_dir.glob("*.csv") if p.name != "summary.csv"]
        assert len(prop_csvs) == len(data["properties"]), \
            f"expected {len(data['properties'])} property CSVs, got {len(prop_csvs)}"
        text = (out_dir / "summary.csv").read_text(encoding="utf-8")
        assert "RECOMMEND-ONLY" in text, "recommend-only note missing from summary"
    print("  ok: CSV fallback (summary + one CSV per property, recommend-only note present)")


def test_partial_payload_does_not_crash():
    """A skinny payload (just names) must still produce output."""
    data = {"properties": [{"name": "Only A Name"}, {"name": "Another"}]}
    with tempfile.TemporaryDirectory() as tmp:
        out_dir = bw.build_csv_fallback(data, Path(tmp) / "r.xlsx")
        assert (out_dir / "summary.csv").exists()
    print("  ok: partial payload (names only) handled without crashing")


def test_duplicate_names_get_unique_tabs():
    data = {"properties": [{"name": "Cabin"}, {"name": "Cabin"}, {"name": "Cabin"}]}
    with tempfile.TemporaryDirectory() as tmp:
        out_dir = bw.build_csv_fallback(data, Path(tmp) / "r.xlsx")
        csvs = sorted(p.name for p in out_dir.glob("*.csv") if p.name != "summary.csv")
        assert len(csvs) == 3, f"duplicate names collapsed: {csvs}"
    print(f"  ok: duplicate property names disambiguated -> {csvs}")


def test_xlsx_if_available():
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  skip: openpyxl not installed — xlsx path not exercised here "
              "(skill installs it at runtime; CSV fallback is what runs without it)")
        return
    data = _load_sample()
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "Revenue-Report-test.xlsx"
        written = bw.build_xlsx(data, out)
        assert written.exists(), "xlsx not written"
        wb = load_workbook(written)
        names = wb.sheetnames
        assert names[0] == "Summary", f"first tab must be Summary, got {names[0]}"
        assert len(names) == 1 + len(data["properties"]), \
            f"expected 1 summary + {len(data['properties'])} property tabs, got {names}"
        for p in data["properties"]:
            assert any(p["name"][:20] in n for n in names), f"missing tab for {p['name']}"
    print(f"  ok: xlsx built — tabs = {names}")


def main():
    print("build_workbook smoke test:")
    failures = 0
    for fn in (test_csv_fallback, test_partial_payload_does_not_crash,
               test_duplicate_names_get_unique_tabs, test_xlsx_if_available):
        try:
            fn()
        except AssertionError as e:
            failures += 1
            print(f"  FAIL: {fn.__name__}: {e}")
    if failures:
        print(f"\n{failures} failure(s).")
        return 1
    print("\nall checks passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
